# coding=utf-8
import openpyxl
import os
import sys

PRIVATE = 'private'
SERVICE = 'service'
CONTROLLER = 'controller'
CHAR_TAB = '\t'
CHAR_ENTER = '\n'
CHAT_SPACE = ' '


def is_config_exists():
    is_exists = os.path.exists(os.path.join(os.getcwd(), 'config.ini'))
    if not is_exists:
        print('Error: config.ini is not found')
        return False
    return True


def read_excel_name():
    if not is_config_exists():
        return ''
    config_path = os.path.join(os.getcwd(), 'config.ini')
    config_file = open(config_path, 'r')
    excel_name = config_file.readline()
    if excel_name == '':
        print('Error: the name of excel is not found. please fill in your name of excel in the first line')
        return ''
    return excel_name


def is_excel_exists():
    excel_name = read_excel_name()
    if excel_name == '':
        return False
    excel_path = os.path.join(os.getcwd(), excel_name).replace('\n', '').strip()
    is_excel_path_exists = os.path.exists(excel_path)
    if not is_excel_path_exists:
        print('Error: the excel is not found. Please chick the excel is if exists')
        return False
    return True;


def get_excel_name():
    if not is_excel_exists():
        return ''
    excel_name = read_excel_name()
    return excel_name


def creat_result_path():
    result_file_dir = os.path.join(os.getcwd() + os.path.sep + 'result' + os.path.sep + 'sql')
    is_exists = os.path.exists(result_file_dir)
    if not is_exists:
        os.makedirs(result_file_dir)
    return result_file_dir


def str_init(string):
    if string is None:
        return ''
    else:
        return string


def create_entities(java_entities_name, sheet, annocation, package):
    maxRow = sheet.max_row

    create_result = 'package ' + package + '.entities;\n\n'
    create_result += annocation.strip() + '\n'
    print(create_result)

    create_result += 'public class ' + java_entities_name + '{\n';
    for i in range(5, maxRow):
        property_name = sheet.cell(row=i, column=1).value
        property_type = sheet.cell(row=i, column=2).value

        create_result += '\tprivate '
        create_result += property_type + ' '
        create_result += property_name + ''
        create_result += ';\n'

    create_result += '}\n'
    result_file_dir = creat_result_path()
    current_file_path = os.path.join(result_file_dir, java_entities_name + '.java')

    file = open(current_file_path, 'w')

    file.write(create_result)

    file.close

    return True


def create_controller(java_entities_name, sheet, annocation, package):

    create_result = 'package ' + package + '.controller;\n\n'
    create_result += annocation.strip() + '\n'
    print(create_result)

    create_result += 'public class ' + java_entities_name + 'Controller{\n';
    create_result += '\t@Autowired\n'
    create_result += CHAR_TAB + PRIVATE + java_entities_name.capitalize() + SERVICE
    create_result += java_entities_name + SERVICE + CHAR_ENTER

    create_result += '}\n'
    result_file_dir = creat_result_path()
    current_file_path = os.path.join(result_file_dir, java_entities_name + '.java')

    file = open(current_file_path, 'w')

    file.write(create_result)

    file.close

def create_java_code():
    excel_name = get_excel_name()
    if excel_name == '':
        return False
    sheet = openpyxl.load_workbook(excel_name).get_sheet_by_name('Sheet1')

    java_entities_name = sheet['A2'].value
    if java_entities_name is None:
        print('Error: the name of java_entities_name must be not null')

    java_package_name = sheet['B2'].value
    if java_package_name is None:
        print('Error: the name of java_package_name must be not null')

    controller_annocation = str_init(sheet['C2'].value)
    service_annocation = str_init(sheet['D2'].value)
    repository_annocation = str_init(sheet['E2'].value)
    entities_annocation = str_init(sheet['F2'].value)
    vo_annocation = str_init(sheet['G2'].value)

    create_entities(java_entities_name, sheet, entities_annocation, java_package_name)



def fun():
    excel_name = get_excel_name()
    if excel_name == '':
        return False
    sheet = openpyxl.load_workbook(excel_name).get_sheet_by_name('Sheet1')

    table_name = sheet['A2'].value
    if table_name is None:
        object_name = 'uPig'
    table_primary_key = sheet['B2'].value
    if table_primary_key is None:
        result_name = 'uDog'

    maxRow = sheet.max_row

    create_table_sql = 'create table if not exists `' + table_name + '`( \n';
    for i in range(5, maxRow):
        item_name = sheet.cell(row=i, column=1).value
        item_type = sheet.cell(row=i, column=2).value
        item_length = sheet.cell(row=i, column=3).value
        item_default_value = sheet.cell(row=i, column=4).value
        item_not_null = sheet.cell(row=i, column=5).value
        item_auto_increment = sheet.cell(row=i, column=6).value
        item_comment = sheet.cell(row=i, column=7).value


        create_table_sql += '\t`' + item_name + '` '
        create_table_sql += item_type + '('
        create_table_sql += str(item_length) + ') '
        if item_default_value is not None:
            create_table_sql += 'default ' + str(item_default_value) + ' '
        if item_not_null is not None:
            create_table_sql += str(item_not_null) + ' '
        if item_auto_increment is not None:
            create_table_sql += item_auto_increment + ' '
        if item_comment is not None:
            create_table_sql += 'comment "' + str(item_comment) + '"'
        create_table_sql += ',\n'

    create_table_sql += '\tprimary key (' + table_primary_key + ')\n)'
    create_table_sql += 'ENGINE = InnoDB \n DEFAULT CHARSET = utf8 \n COLLATE = utf8_bin; \n'

    result_file_dir = creat_result_path()
    current_file_path = os.path.join(result_file_dir, excel_name.replace('.xlsx', '').strip() + '.sql')

    file = open(current_file_path, 'w')

    file.write(create_table_sql)

    file.close

    return True


def main():
    create_java_code()
    os.system('pause')


if __name__ == '__main__':
    sys.exit(main())
