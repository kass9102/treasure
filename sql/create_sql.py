# coding=utf-8
import openpyxl
import os
import sys


def is_config_exists():
    is_exists = os.path.exists(os.path.join(os.getcwd(), 'config.ini'))
    if not is_exists:
        print('Error: config.ini is not found')
        return False
    return True


def read_excel_name():
    if not is_config_exists():
        return ''
    config_path = os.path.join(os.getcwd(), 'config.ini')
    config_file = open(config_path, 'r')
    excel_name = config_file.readline()
    if excel_name == '':
        print('Error: the name of excel is not found. please fill in your name of excel in the first line')
        return ''
    return excel_name


def is_excel_exists():
    excel_name = read_excel_name()
    if excel_name == '':
        return False
    excel_path = os.path.join(os.getcwd(), excel_name).replace('\n', '').strip()
    is_excel_path_exists = os.path.exists(excel_path)
    if not is_excel_path_exists:
        print('Error: the excel is not found. Please chick the excel is if exists')
        return False
    return True;


def get_excel_name():
    if not is_excel_exists():
        return ''
    excel_name = read_excel_name()
    return excel_name


def creat_result_path():
    result_file_dir = os.path.join(os.getcwd() + os.path.sep + 'result' + os.path.sep + 'sql')
    is_exists = os.path.exists(result_file_dir)
    if not is_exists:
        os.makedirs(result_file_dir)
    return result_file_dir


def fun():
    excel_name = get_excel_name()
    if excel_name == '':
        return False
    sheet = openpyxl.load_workbook(excel_name).get_sheet_by_name('Sheet1')

    table_name = sheet['A2'].value
    if table_name is None:
        object_name = 'uPig'
    table_primary_key = sheet['B2'].value
    if table_primary_key is None:
        result_name = 'uDog'

    maxRow = sheet.max_row

    create_table_sql = 'create table if not exists `' + table_name + '`( \n';
    for i in range(5, maxRow):
        item_name = sheet.cell(row=i, column=1).value
        item_type = sheet.cell(row=i, column=2).value
        item_length = sheet.cell(row=i, column=3).value
        item_default_value = sheet.cell(row=i, column=4).value
        item_not_null = sheet.cell(row=i, column=5).value
        item_auto_increment = sheet.cell(row=i, column=6).value
        item_comment = sheet.cell(row=i, column=7).value


        create_table_sql += '\t`' + item_name + '` '
        create_table_sql += item_type + '('
        create_table_sql += str(item_length) + ') '
        if item_default_value is not None:
            create_table_sql += 'default ' + str(item_default_value) + ' '
        if item_not_null is not None:
            create_table_sql += str(item_not_null) + ' '
        if item_auto_increment is not None:
            create_table_sql += item_auto_increment + ' '
        if item_comment is not None:
            create_table_sql += 'comment "' + str(item_comment) + '"'
        create_table_sql += ',\n'

    create_table_sql += '\tprimary key (' + table_primary_key + ')\n)'
    create_table_sql += 'ENGINE = InnoDB \n DEFAULT CHARSET = utf8 \n COLLATE = utf8_bin; \n'

    result_file_dir = creat_result_path()
    current_file_path = os.path.join(result_file_dir, excel_name.replace('.xlsx', '').strip() + '.sql')

    file = open(current_file_path, 'w')

    file.write(create_table_sql)

    file.close

    return True


def main():
    fun()
    os.system('pause')


if __name__ == '__main__':
    sys.exit(main())
